"""
This script generates personalized emails from templates for the purpose of collecting asset information
based on login and machine details stored in a spreadsheet. 
"""


import pandas as pd
from openpyxl import load_workbook

dup_email_template = """Hello {name},

This is [insert name] from IT. I am working on a project to update our machine asset records and am seeking information on the {type} with machine name {machine}.

If available, I will need two pieces of information:

1. Exact location, down to cube or office if possible (e.g., ...)
2. User of the computer
      - If you are the primary/personal user of this device, please confirm this
      - If you know the primary/personal user of this device, please list the name
      - If this is a shared computer, give the group it is shared by, e.g., department or office, and if there is a standard login name used

If you have any questions about this process, please let me know.

Thank you in advance for your assistance!"""

email_template = """Hello {name},

This is [insert name] from IT. I am working on a project to update our machine asset records and am seeking information on the {type} with machine name {machine}.

If available, please fill out this short form that asks a couple quick questions clarifying the location and user of the computer: [insert form link]


If you have any questions about this process, please let me know.

Thank you in advance for your assistance!"""

df = pd.read_excel('data.xlsx')
wb = load_workbook('data.xlsx')
ws = wb.active  # or wb['SheetName']

final_email_col = df.columns.get_loc('Email') + 1  # openpyxl is 1-indexed
dup_col = df.columns.get_loc('User seen previously?') + 1  # openpyxl is 1-indexed

users = {}


def generate_emails():
    start = 2 # lower bound
    # end = 110 # upper bound

    for index, row in df.iterrows():
        if index < start - 2:
            continue
        # if index > end - 1: 
        #     break

        # if row['Missing info?'] == 1 or row['Sent?'] == 1 or row['Dont email/complete?']== 1:
        #     continue

        logon_name = row['Last login']
        logon_name_raw = row['Last login']

        logon_name = str(logon_name).replace(".", " ") # john.smith -> john smith
        name_parts = logon_name.split() # ["john", "smith"]
        logon_name = " ".join(name_parts[:-1]) # john smith -> john
        logon_name = logon_name.capitalize() # john -> john
        

        machine_name = row['Machine name']
        machine_type = "laptop" if machine_name[-1] == 'L' else 'desktop'

        email_final = email_template.format(name = logon_name, type = machine_type, machine = machine_name)
        # print(email_final)


        # find duplicate users
        if logon_name_raw not in users:
            users[logon_name_raw] = 1
        else: # if we do find a dup user, we send them the traditional email so as to traceback the original machine name
            email_final = dup_email_template.format(name = logon_name, type = machine_type, machine = machine_name)
            # ws.cell(row=index+2, column=final_email_col, value=email_final)
            # ws.cell(row=index+2, column=dup_col, value=1)
            users[logon_name_raw] += 1
            # print(f'{logon_name_raw}: {users[logon_name_raw]}')
            continue


        ws.cell(row=index+2, column=final_email_col, value=email_final)
    # print(sorted(users, key=users.get, reverse=True))
    # print(sorted(users.items(), key=lambda x: x[1], reverse=True))

    wb.save('data.xlsx')

generate_emails()

